import openpyxl as xl
import re, os


def fill_info(filename, row):
    values = re.findall(ur'([\w+,\u4e00-\u9fa5\uff00-\uffef]+)\s*\uff08([\S+,\u4e00-\u9fa5\uff00-\uffef]+)\uff09\s*-\s*([\S+,\u4e00-\u9fa5\uff00-\uffef]+).jpg', filename)
    if len(values) == 0:
        print 'line match failed of: ' + filename
        return False
    else:
        ws.cell('A' + str(row)).value = values[0][0]
        ws.cell('B' + str(row)).value = values[0][1]
        ws.cell('C' + str(row)).value = values[0][2]
        return True

FILE_NAME = 'data.xlsx'
wb = xl.load_workbook(filename=FILE_NAME)
ws = wb.get_active_sheet()
row = 2

for fn0 in os.listdir("photo"):
    fn = fn0.decode("utf8")
    if fill_info(fn, row):
        row += 1
    else:
        os.remove("photo/" + fn0)

wb.save(filename='new-' + FILE_NAME)