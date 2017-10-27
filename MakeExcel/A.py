import openpyxl as xl
import re


def add_up(cell_value):
    values = re.findall(r'(\d+)-(\d+)-(\d+)', cell_value)
    if len(values) == 0:
        print 'line match failed of: ' + cell_value
        return 0
    else:
        return sum([int(value) for value in values[0]])


FILE_NAME = 'data.xlsx'
wb = xl.load_workbook(filename=FILE_NAME)
ws = wb.get_active_sheet()
i = 2

for row in ws.rows:
    cell_value = ws.cell('A' + str(i)).value
    if cell_value is not None:
        ws.cell('F' + str(i)).value = add_up(cell_value)
    i += 1

wb.save(filename= 'new-' + FILE_NAME)
